import os
import re
import openpyxl
import threading
from queue import Queue
from tqdm import tqdm
import tkinter as tk
from tkinter import filedialog
# run tree --charset ascii -flashuR /path/to/directory/ > file.txt in the command prompt to generate the tree.txt file
def parse_tree_file(filename, queue):
    with open(filename, 'r', encoding='ascii') as file:
        total_lines = sum(1 for _ in file)
        file.seek(0)  # Reset file pointer to the start
        for line_number, line in tqdm(enumerate(file, start=1), total=total_lines, desc="Parsing file"):
            match = re.search(r"\[(.*?)\]\s*(.*)", line.strip())
            if match:
                size, path = match.groups()
                indent = len(line) - len(line.lstrip(' |-'))  # Count leading spaces or dashes
                level = indent // 4  # Assuming every 4 spaces or dashes denote a new level
                queue.put((level + 1, size, path))  # level + 1 to make Excel column index
            else:
                queue.put(("error", f"No match found on line {line_number}"))
    queue.put(None)  # Signal that parsing is complete

def create_excel(queue, output_filename):
    file_index = 1
    wb = openpyxl.Workbook()
    ws = wb.active
    row_count = 1  # Start row_count at 1

    while True:
        entry = queue.get()
        if entry is None:  # Check for end signal
            break
        if entry[0] == "error":
            print(entry[1])  # Print error message
            continue
        col, size, path = entry
        if row_count >= 1048576:  # Check for row limit
            wb.save(f"{output_filename}_{file_index}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            row_count = 1
            file_index += 1
        ws.cell(row=row_count, column=col, value=f"[{size}] {path}")
        row_count += 1

    wb.save(f"{output_filename}_{file_index}.xlsx")
    print(f"Excel files saved starting with {output_filename}")

def process_file(filename):
    queue = Queue()
    parser_thread = threading.Thread(target=parse_tree_file, args=(filename, queue))
    excel_thread = threading.Thread(target=create_excel, args=(queue, os.path.splitext(os.path.basename(filename))[0]))

    parser_thread.start()
    excel_thread.start()

    parser_thread.join()
    excel_thread.join()

    print("All operations completed successfully.")

def main():
    root = tk.Tk()
    root.withdraw()  # Hides the tkinter window since we only want the file dialog
    filename = filedialog.askopenfilename(
        title="Select a file",
        filetypes=(("Text files", "*.txt"), ("All files", "*.*"))
    )
    if not filename:
        print("No file selected.")
        return  # Exit if no file is selected
    
    print(f"Starting processing of file: {filename}")
    process_file(filename)

# Run the main script
if __name__ == "__main__":
    main()
